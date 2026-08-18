"""
Export routes — CSV, Excel (.xlsx) and PDF download of the ledger.

GET /export/transactions?format=csv|xlsx|pdf

The three builders below are ported from DPL_project/backend/routers/export.py
essentially unchanged: they are already generic over
(col_names, col_display, col_types, rows), so nothing about DPL's master table
leaked into them. Only the data source and the document title differ -- this
exports the transactions ledger with its master-table names resolved, rather
than a dynamic column set.
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from fpdf import FPDF

from database import company_connection
from routers.auth import get_current_schema, get_current_user
from services import custom_fields, scoping

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])

_NUMERIC_TYPES = {"numeric", "real", "double precision", "integer", "bigint"}
_INTEGER_TYPES = {"integer", "bigint", "smallint"}
_DATE_TYPES = {"date", "timestamp", "timestamp without time zone",
               "timestamp with time zone"}

# The five master labels every ledger row resolves to, as
# (SQL source, output name, header). These are joins, not columns, so they have
# no fieldmap row to take a name from -- unlike every statement field, which
# does. Nothing else about the export shape is written down here: the columns,
# their order and their headers all come from the live table via
# custom_fields.data_columns().
_MASTER_LABELS = [
    ("bm.bank_name", "bank_name", "Bank"),
    ("p.name", "project_name", "Project"),
    ("h.name", "head_name", "Head"),
    ("rh.name", "rera_head_name", "RERA Head"),
    ("ih.name", "idw_head_name", "IDW Head"),
    ("ben.name", "beneficiary_name", "Beneficiary"),
]


# --- helpers ------------------------------------------------------------------

def _is_numeric(col_type: str) -> bool:
    return (col_type or "").lower() in _NUMERIC_TYPES


def _is_integer(col_type: str) -> bool:
    """Whole-number column — money formatting would render an id as "41.00"."""
    return (col_type or "").lower() in _INTEGER_TYPES


def _is_date(col_type: str) -> bool:
    return (col_type or "").lower() in _DATE_TYPES


def _as_date(val):
    """Coerce a cell value to date/datetime for Excel, or None if it isn't one."""
    if isinstance(val, (datetime, date)):
        return val
    s = str(val).strip()
    for parse in (date.fromisoformat, datetime.fromisoformat):
        try:
            return parse(s)
        except (ValueError, TypeError):
            continue
    return None


async def _fetch_ledger(user: dict, project_id, head_id, date_from, date_to):
    """The rows to export, with the same filters /transactions accepts.

    Debit and credit are split back into two columns here. The ledger stores one
    amount plus a CR/DR marker, but an accountant reading an exported statement
    expects the two-column bank layout, and Excel totals are per column.

    Project scoping is applied here rather than at the endpoint so that an
    export can never contain a row the same user could not see on screen.
    """
    filters, params, idx = ["1=1"], [], 1
    for value, clause in (
        (project_id, "t.project_id = ${}"),
        (head_id, "t.head_id = ${}"),
    ):
        if value is not None:
            filters.append(clause.format(idx))
            params.append(value)
            idx += 1

    async with company_connection(user["schema"]) as conn:
        # Everything about the shape of this export is read from the live
        # tables: which columns exist, what they are called, which one holds
        # the date. Nothing below names a statement field.
        columns = await custom_fields.data_columns(conn)
        date_col = await custom_fields.date_column(conn)

        for value, op in ((date_from, ">="), (date_to, "<=")):
            if value is not None and date_col:
                filters.append(f"t.{date_col} {op} ${idx}")
                params.append(value)
                idx += 1

        scope = await scoping.visible_project_ids(conn, user)
        if scoping.scope_is_empty(scope):
            return [], []
        # include_unassigned=False: the "unfiled rows belong to everyone" rule
        # exists so a fresh import can be classified, which only concerns
        # staging. A row that reached the ledger with no project is filed data
        # nobody's project owns, and only admins see it.
        clause, scope_params, idx = scoping.project_filter(
            scope, "t.project_id", idx, include_unassigned=False
        )
        if clause:
            filters.append(clause)
            params.extend(scope_params)

        # The company's own columns, then the master names each id resolves to.
        # When amount + credit_debit are among them the pair is expanded into
        # the two-column Debit/Credit layout an accountant expects; a company
        # whose fieldmap already has separate debit and credit columns exports
        # those directly and needs no such expansion.
        select_parts, out_columns = [], []
        for c in columns:
            if c["name"] == "credit_debit":
                continue
            if c["name"] == "amount":
                select_parts += [
                    "CASE WHEN t.credit_debit = 'DR' THEN t.amount END AS debit",
                    "CASE WHEN t.credit_debit = 'CR' THEN t.amount END AS credit",
                ]
                # Both headers are built from the amount column's own live
                # label, so renaming that field in the fieldmap renames these
                # too. They read "Debit"/"Credit" from a literal here, which
                # made them the one pair of headers in the export that the
                # fieldmap could not reach. DR and CR are the values stored in
                # credit_debit, not names chosen here.
                money = c["displayname"] or c["name"]
                out_columns += [
                    {"name": "debit", "displayname": f"{money} (DR)", "type": "numeric"},
                    {"name": "credit", "displayname": f"{money} (CR)", "type": "numeric"},
                ]
                continue
            select_parts.append(f"t.{c['name']}")
            out_columns.append(c)

        select_parts += [f"{src} AS {alias}" for src, alias, _ in _MASTER_LABELS]
        out_columns += [
            {"name": alias, "displayname": label, "type": "text"}
            for _, alias, label in _MASTER_LABELS
        ]

        rows = await conn.fetch(
            f"""
            SELECT {", ".join(select_parts)}
            FROM transactions t
            LEFT JOIN temp_trans        ti  ON ti.id  = t.temp_trans_id
            LEFT JOIN bank_master        bm  ON bm.id  = t.bank_id
            LEFT JOIN projects           p   ON p.id   = t.project_id
            LEFT JOIN head_master        h   ON h.id   = t.head_id
            LEFT JOIN rera_head_master   rh  ON rh.id  = t.rera_head_id
            LEFT JOIN idw_head_master    ih  ON ih.id  = t.idw_head_id
            LEFT JOIN beneficiary_master ben ON ben.id = t.beneficiary_id
            WHERE {" AND ".join(filters)}
            ORDER BY {f"t.{date_col}," if date_col else ""} t.id
            """,
            *params,
        )
    return out_columns, [dict(r) for r in rows]


# --- CSV ----------------------------------------------------------------------

def _build_csv(col_names, col_display, col_types, rows) -> tuple[bytes, str, str]:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(col_display)

    for row in rows:
        out = []
        for c in col_names:
            val = row.get(c, "")
            out.append("" if val is None else str(val))
        writer.writerow(out)

    # utf-8-sig, not utf-8: Excel on Windows reads a BOM-less UTF-8 CSV as
    # cp1252 and mangles any non-ASCII name.
    return buf.getvalue().encode("utf-8-sig"), "text/csv", "csv"


# --- Excel --------------------------------------------------------------------

def _build_xlsx(col_names, col_display, col_types, rows) -> tuple[bytes, str, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    hdr_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
    for ci, label in enumerate(col_display, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill

    # Cells are typed from the declared column type so Excel gets real dates and
    # numbers instead of text — sorting, date filters and number formatting all
    # depend on the cell type, not on how the value looks.
    # Widths are measured from the RENDERED text, not the raw value: "#,##0.00"
    # turns 4000000.0 into "4,000,000.00", three characters wider than str()
    # reports, and a column sized from str() renders the cell as "#####".
    widths = [len(str(lbl)) for lbl in col_display]

    for ri, row in enumerate(rows, 2):
        for ci, c in enumerate(col_names, 1):
            val = row.get(c, "")
            ctype = col_types.get(c, "")
            cell = ws.cell(row=ri, column=ci)
            shown = ""
            if val is None or val == "":
                cell.value = None
            elif _is_date(ctype):
                parsed = _as_date(val)
                if parsed is None:
                    cell.value = shown = str(val)
                else:
                    cell.value = parsed
                    cell.number_format = "yyyy-mm-dd"
                    shown = "0000-00-00"
            elif _is_integer(ctype):
                s = str(val).replace(",", "").strip()
                try:
                    cell.value = int(s) if s.lstrip("-").isdigit() else int(float(s))
                    cell.number_format = "0"
                    shown = str(cell.value)
                except (ValueError, TypeError):
                    cell.value = shown = str(val)
            elif _is_numeric(ctype):
                try:
                    cell.value = float(str(val).replace(",", ""))
                    cell.number_format = "#,##0.00"
                    shown = f"{cell.value:,.2f}"
                except (ValueError, TypeError):
                    cell.value = shown = str(val)
            else:
                cell.value = shown = str(val)
            widths[ci - 1] = max(widths[ci - 1], len(shown))

    # Pad by 2 for cell margins; cap at 50 so one long narration can't stretch
    # the sheet off-screen.
    for ci in range(1, len(col_names) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(widths[ci - 1] + 2, 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return (buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx")


# --- PDF ----------------------------------------------------------------------

class _ExportPDF(FPDF):
    """Table-based PDF export — fixed row heights, no wrapping chaos."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.row_count = 0
        self.doc_title = "Transactions Export"

    def header(self) -> None:
        self.set_font("Helvetica", "B", 13)
        self.cell(0, 7, self.doc_title, new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 5,
                  f"Generated: {date.today().isoformat()}     Rows: {self.row_count}",
                  new_x="LMARGIN", new_y="NEXT")
        self.ln(1)

    def footer(self) -> None:
        self.set_y(-10)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def _build_pdf(col_names, col_display, col_types, rows, title="Transactions Export"):
    pdf = _ExportPDF(orientation="L", unit="mm", format="A4")
    pdf.row_count = len(rows)
    pdf.doc_title = title
    # Page breaks are handled manually so every page repeats the header row.
    # Leaving fpdf's automatic break on as well splits a row across pages and
    # re-enters header() mid-row with a different font size.
    pdf.set_auto_page_break(auto=False, margin=12)

    usable_w = pdf.w - pdf.l_margin - pdf.r_margin  # landscape A4 ~ 277 mm
    n = len(col_names)

    # Widths come from actual content, never a name-keyed table. Weight each
    # column by its widest sampled cell, clamp so one long narration can't
    # starve the rest, then normalize so the table ends on the right margin.
    sample = rows[:200]
    weights = []
    for i, c in enumerate(col_names):
        longest = len(str(col_display[i]))
        for r in sample:
            v = r.get(c)
            if v is not None:
                longest = max(longest, len(str(v)))
        weights.append(float(min(max(longest, 6), 70)))

    total_wt = sum(weights) or 1.0
    col_w = [max(usable_w * (w / total_wt), 8.0) for w in weights]
    scale = usable_w / sum(col_w)  # the 8 mm floor can push the sum over
    col_w = [w * scale for w in col_w]

    row_h, line_h, pad = 5.5, 3.2, 0.5

    def _fit(text: str, w: float) -> str:
        """Trim text to what fits on one line of w mm at the current font."""
        avail = w - 2 * pad
        if avail <= 0:
            return ""
        if pdf.get_string_width(text) <= avail:
            return text
        lo, hi = 0, len(text)
        while lo < hi:
            mid = (lo + hi + 1) // 2
            if pdf.get_string_width(text[:mid] + "...") <= avail:
                lo = mid
            else:
                hi = mid - 1
        return text[:lo] + "..." if lo else ""

    def _wrap(text: str, w: float) -> list:
        """Split text into the lines fpdf will lay out in a w mm cell."""
        if not text:
            return [""]
        return list(pdf.multi_cell(w - 2 * pad, line_h, text, dry_run=True,
                                   output="LINES")) or [""]

    def _header_row() -> None:
        # Must set its own font: header() leaves it at 13/9 pt, which clips the
        # labels on the first page only.
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        for label, w in zip(col_display, col_w):
            pdf.cell(w, row_h, _fit(str(label), w), border=1, align="L", fill=True)
        pdf.ln(row_h)

    def _draw_row(row: dict) -> None:
        pdf.set_font("Helvetica", "", 7)
        cells = []
        for c, w in zip(col_names, col_w):
            val = row.get(c, "")
            val = "" if val is None else str(val)
            cells.append(_wrap(val, w))

        h = max(row_h, max(len(ls) for ls in cells) * line_h + 1.4)

        if pdf.get_y() + h > pdf.h - pdf.b_margin:
            pdf.add_page()
            _header_row()
            pdf.set_font("Helvetica", "", 7)

        x, y = pdf.l_margin, pdf.get_y()
        for c, w, lines in zip(col_names, col_w, cells):
            pdf.rect(x, y, w, h)
            pdf.set_xy(x + pad, y + 0.7)
            pdf.multi_cell(w - 2 * pad, line_h, "\n".join(lines), border=0,
                           align="R" if _is_numeric(col_types.get(c, "")) else "L")
            x += w
        pdf.set_xy(pdf.l_margin, y + h)

    if n == 0 or not rows:
        pdf.add_page()
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "No data to export.", new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.add_page()
        _header_row()
        for row in rows:
            _draw_row(row)

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue(), "application/pdf", "pdf"


# --- endpoint -------------------------------------------------------------------

@router.get("/transactions")
async def export_transactions(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    project_id: int = Query(None),
    head_id: int = Query(None),
    date_from: date = Query(None, description="YYYY-MM-DD"),
    date_to: date = Query(None, description="YYYY-MM-DD"),
    user: dict = Depends(get_current_user),
):
    """
    Download the ledger as CSV, Excel or PDF.

    Accepts the same filters as GET /transactions, so what you export is what
    you were looking at.
    """
    columns, rows = await _fetch_ledger(user, project_id, head_id, date_from, date_to)

    col_names = [c["name"] for c in columns]
    col_display = [c["displayname"] for c in columns]
    col_types = {c["name"]: c["type"] for c in columns}

    filename = f"transactions_{date.today().isoformat()}"

    try:
        if format == "csv":
            content, media_type, ext = _build_csv(col_names, col_display, col_types, rows)
        elif format == "xlsx":
            content, media_type, ext = _build_xlsx(col_names, col_display, col_types, rows)
        else:
            content, media_type, ext = _build_pdf(col_names, col_display, col_types, rows)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Export failed (%s)", format)
        raise HTTPException(status_code=500, detail=f"Export failed: {exc}")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}.{ext}"'},
    )
