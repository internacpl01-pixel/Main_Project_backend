"""
Spreadsheet and CSV bank-statement import.

Both formats reduce to a grid of strings, so they share one pipeline from there:
the same fieldmap, the same header detection, the same normalization and the
same batch write as the PDF path.

Where this deliberately parts company with the PDF path is row assembly, and
that is the whole point of this module.

  A PDF has no rows. It has lines, and one transaction's narration wraps across
  several of them, so parsers._assemble_rows merges any line that does not start
  a new transaction into the one above it. That is correct for a document made
  of lines and catastrophic for one made of records: fed a spreadsheet, it
  merged 232 rows into 1, joining every description, every reference and every
  amount into single fields thousands of characters long.

  A spreadsheet row IS a transaction. There is no wrapping to undo, so
  _assemble_tabular_rows below maps cells to fields one row at a time and never
  merges anything. Nothing about a statement can make it produce fewer rows than
  the file has.

The other half is sheets. A workbook is not one statement — the file this was
rebuilt against holds eight bank accounts on eight sheets, plus a pivot table
and two beneficiary lists. Reading openpyxl's `active` sheet imported one of
them and silently ignored the rest. Sheets are now inspected, chosen and staged
one batch each, which is the spreadsheet's answer to the PDF page selector.

Replaces services/excel_import.py.
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import re
import time
from datetime import date, datetime

import openpyxl

from import_helpers import compute_fill_rates, normalize_parsed_rows
from parsers import (_build_alias_map, _category_map_from_aliases,
                     _detect_header_row, _extract_document_level_fields)
from services import jobs
from services.fieldmap import get_field_mappings, live_col_types
from services.staging import assert_bank_exists, stage_batch

logger = logging.getLogger(__name__)

# How many rows of a sheet the inspect step shows as a sample. Enough to see
# whether the columns landed where they should, small enough that inspecting a
# ten-sheet workbook is still one quick response.
SAMPLE_ROWS = 5

# Cap on the preview payload, matching the PDF path.
PREVIEW_ROWS = 200


def _cell(value) -> str:
    """One spreadsheet cell as the string the parser should see.

    Dates become ISO. Left as datetimes, str() renders "2026-08-04 00:00:00",
    which the date detectors in parsers.py reject — every row would look undated.

    A whole number loses its ".0". Excel has one numeric type, so every integer
    comes back from openpyxl as a float and str() writes 45563200000264.0 — an
    account number with a decimal point on the end, a reference number that no
    longer matches the bank's, and a digits-only comparison that gains a
    trailing zero. Rendered as an integer it is what the sheet displays.

    Excel error values arrive as their literal text ('#NAME?', '#REF!'). They
    are blanked rather than carried: a broken formula is not data, and '#NAME?'
    in a numeric column parses to nothing anyway while still making the cell
    look filled.

    What this cannot recover is a LEADING zero. Excel dropped it when the cell
    was stored, so '045563200000264' is simply not in the file — which is why
    account numbers are matched with leading zeros ignored. See
    services.staging.account_digits.
    """
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        # Only for values that fit an int exactly. A float big enough to have
        # lost precision is left as it is rather than dressed up as an exact
        # integer it no longer is.
        if abs(value) < 2 ** 53:
            return str(int(value))
    text = str(value).strip()
    if text.startswith("#") and text.endswith(("?", "!")) and len(text) <= 10:
        return ""
    return text


def read_excel_sheets(file_bytes: bytes) -> list[tuple[str, list]]:
    """Every sheet in the workbook, in tab order, as (name, grid).

    Every sheet, not openpyxl's `active` one. Which sheet Excel happened to be
    left on when the file was saved is not a statement of intent, and treating
    it as one is how seven of eight accounts went missing without a word.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        out = []
        for name in wb.sheetnames:
            grid = []
            for row in wb[name].iter_rows(values_only=True):
                cells = [_cell(c) for c in row]
                if any(cells):
                    grid.append(cells)
            out.append((name, grid))
        return out
    finally:
        wb.close()


def _read_excel_rows(file_bytes: bytes) -> list:
    """The active sheet only. Kept for callers that want one grid."""
    sheets = read_excel_sheets(file_bytes)
    return sheets[0][1] if sheets else []


def _read_csv_rows(file_bytes: bytes) -> list:
    """Read a CSV as a grid of strings.

    utf-8-sig strips the BOM Excel writes, which would otherwise glue itself to
    the first header cell and stop "Date" matching its alias. The delimiter is
    sniffed because Indian bank exports use comma, semicolon and tab about
    equally; a failed sniff falls back to comma rather than raising.
    """
    text = file_bytes.decode("utf-8-sig", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = []
    for row in csv.reader(io.StringIO(text, newline=""), dialect):
        cells = [(c or "").strip() for c in row]
        if any(cells):
            rows.append(cells)
    return rows


def _csv_sheets(file_bytes: bytes) -> list[tuple[str, list]]:
    """A CSV is a workbook with one unnamed sheet, so the rest of this module
    does not need to know which format it is looking at."""
    return [("", _read_csv_rows(file_bytes))]


READERS = {
    "excel": (_read_excel_rows, (".xlsx", ".xls")),
    "csv": (_read_csv_rows, (".csv",)),
}

_SHEET_READERS = {"excel": read_excel_sheets, "csv": _csv_sheets}


# ── Row assembly ─────────────────────────────────────────────────────────────

def _norm_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (text or "").lower())


def _is_repeat_header(cells: list, header_cells: list) -> bool:
    """True when this row is the header printed again.

    Long exports repeat the column titles every page-worth of rows. Compared on
    the non-empty cells only, and only when at least three of them line up, so a
    transaction that happens to share one word with a header is not mistaken for
    one.
    """
    hits = 0
    for idx, want in enumerate(header_cells):
        if not want or idx >= len(cells) or not cells[idx]:
            continue
        if _norm_header(cells[idx]) == _norm_header(want):
            hits += 1
    return hits >= 3


def _resolve_collisions(header_cells: list, col_mapping: dict,
                        display_by_field: dict,
                        date_cols: set) -> tuple[dict, list]:
    """Decide which column owns a field when several claim it.

    Two source columns can map to one field. This workbook does it twice: HEAD
    and SUB HEAD both land on one column, and TYPE and TYPE FOR RERA IDW both
    land on another. Left alone, whichever came first won per row, so on the two
    BOM sheets — where TYPE holds 'IMPS' and 'Charges' and TYPE FOR RERA IDW
    holds the actual RERA category — the wrong one took the column.

    The column whose title IS the field's display name wins. That is not a
    tie-break, it is the sheet saying which column it means: 'TYPE FOR RERA IDW'
    is that field, and 'TYPE' merely shares a word with one of its aliases.
    Failing an exact match, the leftmost wins, which is the older behaviour and
    the right answer for a transaction-date / value-date pair.

    The losing columns are dropped for that field rather than used as a fallback
    on rows where the winner is blank. Filling 'TYPE FOR RERA IDW' with 'IMPS'
    on the rows that happen to be empty would put two different meanings in one
    column, and a column that is right nine rows in ten is worse than one that
    is empty — you cannot tell which rows to distrust.

    Returns (owner_by_field, reported) — reported names the winner and losers so
    the fieldmap can be corrected, which is the only real fix.
    """
    owner: dict[str, int] = {}
    reported: list[dict] = []

    by_field: dict[str, list] = {}
    for col_idx, fieldname in col_mapping.items():
        by_field.setdefault(fieldname, []).append(col_idx)

    def title(col_idx: int) -> str:
        return str(header_cells[col_idx]) if col_idx < len(header_cells) else ""

    for fieldname, cols in by_field.items():
        cols.sort()
        if len(cols) == 1:
            owner[fieldname] = cols[0]
            continue

        wanted = _norm_header(display_by_field.get(fieldname, ""))
        exact = [c for c in cols if wanted and _norm_header(title(c)) == wanted]
        win = exact[0] if exact else cols[0]
        owner[fieldname] = win

        # A transaction date beside a value date is not a mapping mistake, it is
        # how statements are printed, so it is resolved silently.
        if not all(c in date_cols for c in cols):
            reported.append({
                "field": fieldname,
                "used": title(win) or fieldname,
                "ignored": [title(c) for c in cols if c != win],
                "matched_display_name": bool(exact),
            })
    return owner, reported


def _assemble_tabular_rows(grid: list, header_idx: int, col_mapping: dict,
                           date_cols: set,
                           display_by_field: dict | None = None) -> tuple[list, list, dict]:
    """One grid row in, one record out. Returns (records, collisions, owner).

    No continuation merging, ever. That is the difference between this and the
    PDF assembler and the reason this module exists: a spreadsheet cell already
    holds its whole value, so a row that looks incomplete is an incomplete row,
    not the tail of the one above it. Fed this workbook, the PDF assembler
    returned 1 row for a 232-row sheet.
    """
    header_cells = grid[header_idx] if 0 <= header_idx < len(grid) else []
    owner, collisions = _resolve_collisions(
        header_cells, col_mapping, display_by_field or {}, date_cols
    )
    # Only the winning column of each field is read at all.
    effective = {col_idx: fn for fn, col_idx in owner.items()}

    records: list[dict] = []
    for row_idx in range(header_idx + 1, len(grid)):
        cells = [str(c).strip() if c else "" for c in grid[row_idx]]
        # "-" and its typographic cousins are empty-cell placeholders.
        cells = ["" if c in ("-", "–", "--") else c for c in cells]
        if not any(cells):
            continue
        if header_cells and _is_repeat_header(cells, header_cells):
            continue

        record = {
            fieldname: cells[col_idx]
            for col_idx, fieldname in effective.items()
            if col_idx < len(cells) and cells[col_idx]
        }
        if record:
            records.append(record)

    return records, collisions, owner


# ── Sheet analysis ───────────────────────────────────────────────────────────

def _date_columns(col_mapping: dict, col_types: dict, cat_by_field: dict) -> set:
    """Which mapped columns hold a date, by live column type then by category."""
    out = set()
    for col_idx, fieldname in col_mapping.items():
        col_type = (col_types.get(fieldname) or "").lower()
        if col_type in ("date", "timestamp without time zone", "timestamp"):
            out.add(col_idx)
        elif cat_by_field.get(fieldname) == "date":
            out.add(col_idx)
    return out


def _display_names(fieldmap_rows: list) -> dict:
    """{fieldname: display name}, used to settle which column owns a field."""
    return {
        r.get("fieldname"): (r.get("displayname") or r.get("fieldname") or "")
        for r in (fieldmap_rows or [])
        if r.get("fieldname")
    }


def analyse_sheet(name: str, grid: list, alias_map: dict, col_types: dict,
                  cat_by_field: dict, fieldmap_rows: list) -> dict:
    """Everything known about one sheet without staging any of it.

    `is_statement` is the question the sheet picker actually needs answered, and
    it is decided on structure rather than on the sheet's name: a statement has
    a date column and at least one money column. That is what separates the
    eight account tabs in this workbook from its pivot table and its two
    beneficiary lists, none of which should be offered as a statement to import.
    """
    info = {
        "name": name,
        "rows": len(grid),
        "header_row": None,
        "headers_detected": {},
        "unmapped_headers": [],
        "data_rows": 0,
        "is_statement": False,
        "reason": "",
        "sample": [],
        "document_fields": {},
        "column_collisions": [],
    }

    if not grid:
        info["reason"] = "The sheet is empty."
        return info

    header_idx, col_mapping = _detect_header_row(grid, alias_map)
    if header_idx < 0 or not col_mapping:
        info["reason"] = ("No column titles on this sheet match a fieldmap "
                          "alias, so it cannot be read as a statement.")
        return info

    header_cells = grid[header_idx]
    info["header_row"] = header_idx + 1        # 1-based, as Excel shows it
    info["headers_detected"] = {
        fieldname: header_cells[col_idx]
        for col_idx, fieldname in col_mapping.items()
        if col_idx < len(header_cells)
    }
    info["unmapped_headers"] = [
        str(cell).strip()
        for col_idx, cell in enumerate(header_cells)
        if str(cell).strip() and col_idx not in col_mapping
    ]

    date_cols = _date_columns(col_mapping, col_types, cat_by_field)
    cats = {cat_by_field.get(fn) for fn in col_mapping.values()}
    has_date = bool(date_cols) or "date" in cats
    has_money = bool(cats & {"withdrawal", "deposits", "balance", "amount"})

    records, collisions, owner = _assemble_tabular_rows(
        grid, header_idx, col_mapping, date_cols, _display_names(fieldmap_rows)
    )
    # Reported from the column that WON, not from whichever the mapping happened
    # to list last. Saying "field_date_1 <- VALUE DATE" when the transaction date
    # is what was read describes the wrong column to someone checking the import.
    info["headers_detected"] = {
        fieldname: (header_cells[col_idx] if col_idx < len(header_cells) else fieldname)
        for fieldname, col_idx in owner.items()
    }
    info["data_rows"] = len(records)
    info["column_collisions"] = collisions

    # Values printed above the table — account number, statement period — belong
    # to every row beneath it.
    if records and header_idx > 0:
        filled = set()
        for r in records:
            filled.update(r.keys())
        info["document_fields"] = _extract_document_level_fields(
            "\n".join(" ".join(r) for r in grid[:header_idx]), fieldmap_rows, filled
        )

    if not has_date:
        info["reason"] = "No date column, so this does not look like a statement."
    elif not has_money:
        info["reason"] = ("No debit, credit or balance column, so this does not "
                          "look like a statement.")
    elif not records:
        info["reason"] = "The columns were found but there are no rows under them."
    else:
        info["is_statement"] = True

    info["sample"] = records[:SAMPLE_ROWS]
    info["_records"] = records
    info["_header_idx"] = header_idx
    info["_col_mapping"] = col_mapping
    return info


async def inspect_tabular(schema: str, file_bytes: bytes, kind: str) -> dict:
    """What is in this workbook, sheet by sheet. Nothing is written.

    This is the spreadsheet's answer to the PDF page selector, and it has to run
    before the user can choose: sheet names alone do not say which tabs are
    statements, how many rows each holds, or whether their columns were
    recognised.
    """
    fieldmap_rows = await get_field_mappings(schema)
    if not fieldmap_rows:
        raise RuntimeError(
            "This company has no fieldmap rows, so no column header can be "
            "recognised. Run: python -m db.migrate upgrade"
        )
    alias_map = _build_alias_map(fieldmap_rows)
    cat_by_field = _category_map_from_aliases(alias_map)
    col_types = live_col_types(fieldmap_rows)

    reader = _SHEET_READERS[kind]
    try:
        sheets = reader(file_bytes)
    except Exception as e:
        raise RuntimeError(f"Failed to read {kind} file: {e}")

    analysed = [
        analyse_sheet(name, grid, alias_map, col_types, cat_by_field, fieldmap_rows)
        for name, grid in sheets
    ]
    public = [{k: v for k, v in a.items() if not k.startswith("_")} for a in analysed]
    return {
        "sheets": public,
        "sheet_count": len(public),
        "statement_sheets": [s["name"] for s in public if s["is_statement"]],
        "total_rows": sum(s["data_rows"] for s in public if s["is_statement"]),
    }


def _select(sheets: list, wanted: str) -> list:
    """The sheets to import: the named ones, or every statement sheet.

    Blank means every sheet that looks like a statement — not every sheet, since
    a pivot table has no transactions to take and would only produce an error
    per tab.
    """
    names = [n.strip() for n in (wanted or "").split(",") if n.strip()]
    if not names:
        return [s for s in sheets if s["is_statement"]]

    by_name = {s["name"]: s for s in sheets}
    chosen = []
    for name in names:
        if name not in by_name:
            raise RuntimeError(
                f"This workbook has no sheet called {name!r}. It has: "
                f"{', '.join(repr(s['name']) for s in sheets)}."
            )
        chosen.append(by_name[name])
    return chosen


# ── Import ───────────────────────────────────────────────────────────────────

async def process_tabular_import(
    schema: str,
    file_bytes: bytes,
    filename: str,
    username: str,
    kind: str,
    bank_id: int | None = None,
    save: bool = False,
    sheets: str = "",
    job_id: str | None = None,
) -> dict:
    """Parse a workbook or CSV and, when save=True, stage each sheet as a batch.

    save=False is a dry run: it returns exactly what would be staged, per sheet,
    so the column mapping can be checked before anything is written.

    sheets names which tabs to take, comma separated. Blank takes every sheet
    that looks like a statement.

    One batch per sheet, because one sheet is one account's statement. Staging
    eight accounts as a single batch would tie them all to one bank id and make
    them undiscardable separately, and the file fingerprint is scoped by sheet
    name so importing another tab of the same workbook is a new import rather
    than a duplicate.
    """
    t_start = time.perf_counter()

    fieldmap_rows = await get_field_mappings(schema)
    if not fieldmap_rows:
        raise RuntimeError(
            "This company has no fieldmap rows, so no column header can be "
            "recognised. Run: python -m db.migrate upgrade"
        )
    alias_map = _build_alias_map(fieldmap_rows)
    cat_by_field = _category_map_from_aliases(alias_map)
    col_types = live_col_types(fieldmap_rows)

    # Checked before the file is read, for the same reason as the PDF path: a
    # bank id that cannot be used is knowable in advance.
    if save:
        await assert_bank_exists(schema, bank_id)

    reader = _SHEET_READERS[kind]
    try:
        grids = reader(file_bytes)
    except Exception as e:
        raise RuntimeError(f"Failed to read {kind} file: {e}")

    if not grids or not any(g for _, g in grids):
        raise RuntimeError("The file has no rows.")

    analysed = [
        analyse_sheet(name, grid, alias_map, col_types, cat_by_field, fieldmap_rows)
        for name, grid in grids
    ]
    chosen = _select(analysed, sheets)

    if not chosen:
        names = ", ".join(repr(a["name"]) for a in analysed) or "none"
        raise RuntimeError(
            "No sheet in this file looks like a bank statement — a statement "
            "needs a date column and a debit, credit or balance column. Sheets "
            f"found: {names}. Pick one explicitly if it should have been read."
        )

    results: list[dict] = []
    all_records: list = []
    all_normalized: list = []
    total_staged = 0
    total_duplicates = 0

    for i, sheet in enumerate(chosen, start=1):
        label = sheet["name"] or filename
        records = sheet["_records"]

        if job_id:
            jobs.start_step(job_id, index=i, total=len(chosen), label=label,
                            units=max(1, len(records)),
                            message=f"Reading {label}")

        # Document-level values (account number, period) apply to every row on
        # the sheet they were printed above.
        for record in records:
            for fieldname, value in (sheet["document_fields"] or {}).items():
                record.setdefault(fieldname, value)

        normalized, norm_stats = normalize_parsed_rows(records, fieldmap_rows)
        all_records.extend(records)
        all_normalized.extend(normalized)

        entry = {
            "sheet": sheet["name"],
            "header_row": sheet["header_row"],
            "parsed": len(records),
            "usable": len(normalized),
            "headers_detected": sheet["headers_detected"],
            "unmapped_headers": sheet["unmapped_headers"],
            "column_collisions": sheet["column_collisions"],
            "document_fields": sheet["document_fields"],
            "fill_rates": compute_fill_rates(records),
            "stats": {**norm_stats, "source": kind, "sheet": sheet["name"]},
            "rows": normalized[:PREVIEW_ROWS],
            "truncated": len(normalized) > PREVIEW_ROWS,
            "batch_id": None,
            "staged": 0,
            "duplicate_rows": 0,
            "error": None,
        }

        if save:
            if not normalized:
                # One unusable sheet does not fail the rest. It is reported on
                # its own line and the other tabs still import — the alternative
                # is an eight-sheet workbook refusing entirely because one tab
                # is a pivot table.
                entry["error"] = (
                    "No usable transaction rows on this sheet. Check the "
                    "columns matched — if none, this bank's column names are "
                    "not in the fieldmap yet."
                )
                logger.warning("[%s] sheet %r: nothing usable", kind, sheet["name"])
            else:
                if job_id:
                    jobs.set_state(job_id, jobs.SAVING,
                                   f"Saving {len(normalized)} rows from {label}")
                sheet_label = f"{filename} [{sheet['name']}]" if sheet["name"] else filename
                try:
                    staged = await stage_batch(
                        schema, file_bytes, sheet_label, username, bank_id,
                        normalized, entry["stats"],
                        # Scoped by sheet so each tab is its own upload. Without
                        # it the second sheet of a workbook collides with the
                        # first on file_hash and is refused as a duplicate.
                        hash_scope=f"sheet={sheet['name']}" if sheet["name"] else "",
                    )
                except Exception as exc:                      # noqa: BLE001
                    entry["error"] = str(exc)
                    logger.warning("[%s] sheet %r failed to stage: %s",
                                   kind, sheet["name"], exc)
                else:
                    entry.update(batch_id=staged["batch_id"],
                                 staged=staged["inserted"],
                                 duplicate_rows=staged["duplicate_rows"])
                    total_staged += staged["inserted"]
                    total_duplicates += staged["duplicate_rows"]

        if job_id:
            jobs.complete_step(job_id, rows=entry["staged"] or entry["usable"])
        results.append(entry)

    t_parse = (time.perf_counter() - t_start) * 1000
    logger.info("[%s] %s: %d sheet(s), parsed=%d usable=%d staged=%d in %.0fms",
                kind, filename, len(chosen), len(all_records),
                len(all_normalized), total_staged, t_parse)

    if save and total_staged == 0:
        # Every sheet failed. The per-sheet reasons are the useful part, so they
        # are carried into the message rather than replaced by a generic one.
        detail = "; ".join(
            f"{r['sheet'] or 'the file'}: {r['error']}" for r in results if r["error"]
        )
        raise RuntimeError(detail or "No usable transaction rows were found.")

    # The first sheet describes the import for anything reading the flat fields.
    # Everything per-sheet is in `sheets`, which is what the result screen shows.
    first = results[0]
    merged_headers: dict = {}
    merged_unmapped: list = []
    for r in results:
        for k, v in (r["headers_detected"] or {}).items():
            merged_headers.setdefault(k, v)
        for h in r["unmapped_headers"]:
            if h not in merged_unmapped:
                merged_unmapped.append(h)

    payload = {
        "saved": bool(save and total_staged),
        # Kept for callers written against the single-batch shape. With several
        # sheets there is no one batch id, so it names the first one staged.
        "batch_id": next((r["batch_id"] for r in results if r["batch_id"]), None),
        "row_count": total_staged if save else len(all_normalized),
        "rows": all_normalized[:PREVIEW_ROWS],
        "truncated": len(all_normalized) > PREVIEW_ROWS,
        "headers_detected": merged_headers,
        "unmapped_headers": merged_unmapped,
        "document_fields": first["document_fields"],
        "fill_rates": compute_fill_rates(all_records),
        "stats": {
            "parsed": len(all_records),
            "usable": len(all_normalized),
            "parse_ms": round(t_parse),
            "headers_detected": merged_headers,
            "unmapped_headers": merged_unmapped,
            "source": kind,
            "sheets_read": [r["sheet"] for r in results],
        },
        "duplicate_rows": total_duplicates,
        # The per-sheet breakdown: one line each, with its own batch id, its own
        # counts and its own failure if it had one.
        "sheets": [{k: v for k, v in r.items() if k != "rows"} for r in results],
        "sheets_available": [
            {k: v for k, v in a.items() if not k.startswith("_") and k != "sample"}
            for a in analysed
        ],
        "sheet_count": len(analysed),
        "sheets_imported": len(chosen),
        "total_ms": round((time.perf_counter() - t_start) * 1000),
    }
    return payload


async def start_tabular_job(**kwargs) -> dict:
    """Begin a workbook import in the background and return its job id.

    Same reasoning as the PDF path: the request finishes immediately, so a host
    that caps request duration has nothing to cut short, and the browser can ask
    how far along the import is instead of watching a spinner. A workbook of
    eight sheets and a few thousand rows is not a three-minute parse, but it is
    long enough that a bar reporting sheet 3 of 8 is worth more than a spinner.
    """
    job_id = jobs.create(
        schema=kwargs["schema"],
        username=kwargs["username"],
        filename=kwargs["filename"],
        # A rough total: refined per sheet by start_step, which is what the bar
        # actually reads.
        total_units=1000,
        total_pages=None,
    )

    async def _runner():
        try:
            payload = await process_tabular_import(job_id=job_id, **kwargs)
            jobs.finish(job_id, payload)
        except Exception as exc:                              # noqa: BLE001
            logger.warning("[%s] job %s failed: %s", kwargs.get("kind"), job_id, exc)
            jobs.fail(job_id, str(exc))

    jobs.attach_task(job_id, asyncio.create_task(_runner()))
    return {"job_id": job_id, "state": jobs.QUEUED}
